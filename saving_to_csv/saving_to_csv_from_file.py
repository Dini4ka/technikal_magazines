from technikal_magazines.all_items_parce import *
from openpyxl.styles import Alignment
from copy import copy
import openpyxl


def saving_to_csv_from_file(file_name,shop_name):

    # Открываем загружаемую страницы
    wb = openpyxl.load_workbook(file_name)
    ws = wb.active

    # Создаём выходную страницу
    wb1 = openpyxl.Workbook()
    ws1 = wb1.active
    ws1.title = 'первая страница'

    # Максимальное кол-во столбоцов и строк в загружаемом файле
    mr = ws.max_row
    mc = ws.max_column

    # Копируем загружаемую страницу
    for i in range(1,mr+1):
        for j in range(1, mc+1):
            c = ws.cell(row=i, column=j)
            ws1.cell(row=i, column=j).font = copy(c.font)
            ws1.cell(row=i, column=j).border = copy(c.border)
            ws1.cell(row=i, column=j).fill = copy(c.fill)
            ws1.cell(row=i, column=j).number_format = copy(c.number_format)
            ws1.cell(row=i, column=j).protection = copy(c.protection)
            ws1.cell(row=i, column=j).alignment = copy(c.alignment)
            ws1.cell(row=i, column=j).value = c.value

    # Добавляем поля с оценками
    for j in range(mc + 1,mc + 6):
        c = ws.cell(row=1, column=2)
        ws1.cell(row=1, column=j).font = copy(c.font)
        ws1.cell(row=1, column=j).border = copy(c.border)
        ws1.cell(row=1, column=j).fill = copy(c.fill)
        ws1.cell(row=1, column=j).number_format = copy(c.number_format)
        ws1.cell(row=1, column=j).protection = copy(c.protection)
        ws1.cell(row=1, column=j).alignment = copy(c.alignment)
        ws1.cell(row=1, column=j).alignment = Alignment(horizontal='center')
        ws1.cell(row=1, column=j).value = j -mc

    # Читаем из файла все файлы, которые надо пропарсить
    links = []

    for row in range(2,mr+1):
        links.append(ws1.cell(row=row, column=mc).value)
    print(links)
    result = ''
    # Выбираем магаин
    if shop_name == 'МВидео':
        what_need_to_write = all_Items_Mvideo()
        what_need_to_write.getting_answers_from_file(links)
        result = what_need_to_write.give_me_result()
    if shop_name == 'Эльдорадо':
        what_need_to_write = El_Dorado()
        what_need_to_write.getting_answers_from_file(links)
        result = what_need_to_write.give_me_result()
    if shop_name == 'ДНС':
        a = DNS()
        a.getting_answers_from_file(links)
        result = a.give_me_result()
    if shop_name == 'Ситилинк':
        a = Citilink()
        a.getting_answers_from_file(links)
        result = a.give_me_result()
    if shop_name == 'Холодильник':
        a = Holodilnik()
        a.getting_answers_from_file(links)
        result = a.give_me_result()
    if shop_name == 'РБТ':
        a = RBT()
        a.getting_answers_from_file(links)
        result = a.give_me_result()
    # Максимальное кол-во столбцов в загружаемом файле
    mc1 = ws1.max_column
    row = 1
    for part in result:
        row +=1
        for key in part:
            reviews = part[key]
            count = 0

        # Вставляем оценки
        for j in range(mc1-mc, mc1+1):
            ws1.cell(row=row, column=j).value = reviews[count]
            ws1.cell(row=row, column=j).alignment = Alignment(horizontal='center', vertical='center')
            count += 1

    # Сохраняем фаил
    wb1.save(r"C:/result_files/" + file_name.split('\\')[-1].split('.')[0]+ '_out.xlsx')
