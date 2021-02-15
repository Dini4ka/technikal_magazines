from technikal_magazines.all_items_parce import *
from copy import copy
from openpyxl.styles import Alignment
import openpyxl

def saving_to_csv(shop_link,shop_name):
    if shop_name == 'МВидео':
        a = all_Items_Mvideo()
        a.getting_pages_links(shop_link)
        print('pages links got')
        a.getting_items_links()
        print('items links got')
        a.getting_reviews()
        print('reviews got')
        items_marks = a.give_me_result()
    if shop_name == 'Эльдорадо':
        a = El_Dorado()
        a.find_items_links(shop_link)
        print('items links got')
        a.get_item_reviews()
        print('reviews got')
        items_marks = a.give_me_result()
    if shop_name == 'ДНС':
        a = DNS()
        a.find_items_links(shop_link)
        print('items links got')
        a.get_item_reviews()
        print('reviews got')
        items_marks = a.give_me_result()
    if shop_name == 'Ситилинк':
        a = Citilink()
        a.getting_pages_links(shop_link)
        print('pages links got')
        a.getting_items_links()
        print('items links got')
        a.getting_reviews()
        print('reviews got')
        items_marks = a.give_me_result()
    if shop_name == 'Холодильник':
        a = Holodilnik()
        a.getting_items_links(shop_link)
        print('items links got')
        a.get_item_reviews()
        print('reviews got')
        items_marks = a.give_me_result()

    wb1 = openpyxl.load_workbook(r'C:\technical_shops\donor.xlsx')
    ws1 = wb1['первая страница']

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'первая страница'

    mc = ws1.max_column
    for i in range(1, mc + 1):
        c = ws1.cell(row=1, column=i)
        ws.cell(row=1, column=i).font = copy(c.font)
        ws.cell(row=1, column=i).border = copy(c.border)
        ws.cell(row=1, column=i).fill = copy(c.fill)
        ws.cell(row=1, column=i).number_format = copy(c.number_format)
        ws.cell(row=1, column=i).protection = copy(c.protection)
        ws.cell(row=1, column=i).alignment = copy(c.alignment)
        ws.cell(row=1, column=i).value = copy(c.value)


    for item_mark in range(len(items_marks)):
        print(str(item_mark) + ' string is ready')
        magazin = shop_name
        if shop_name == 'Ситилинк':
            brand = shop_link.split('/')[-2]
        else:
            brand = shop_link.split('=')[len(shop_link.split('='))-1]
        link = ''
        one_star = ''
        two_star = ''
        three_star = ''
        four_star = ''
        five_star = ''
        for key in items_marks[item_mark]:
            link = key
            one_star = items_marks[item_mark][key][0]
            two_star = items_marks[item_mark][key][1]
            three_star = items_marks[item_mark][key][2]
            four_star = items_marks[item_mark][key][3]
            five_star = items_marks[item_mark][key][4]

        row = item_mark + 2

        cell = ws.cell(row=row, column=1)
        cell.value = magazin
        cell.alignment = Alignment(horizontal='center')
        cell = ws.cell(row=row, column=2)
        cell.value = brand
        cell.alignment = Alignment(horizontal='center')
        cell = ws.cell(row=row, column=3)
        cell.value = link
        cell.alignment = Alignment(horizontal='center')
        cell = ws.cell(row=row, column=4)
        cell.value = one_star
        cell.alignment = Alignment(horizontal='center')
        cell = ws.cell(row=row, column=5)
        cell.value = two_star
        cell.alignment = Alignment(horizontal='center')
        cell = ws.cell(row=row, column=6)
        cell.value = three_star
        cell.alignment = Alignment(horizontal='center')
        cell = ws.cell(row=row, column=7)
        cell.value = four_star
        cell.alignment = Alignment(horizontal='center')
        cell = ws.cell(row=row, column=8)
        cell.value = five_star
        cell.alignment = Alignment(horizontal='center')
    name = shop_name + '_' + brand + '_out.xlsx'
    wb.save(r"C:/result_files/" + name)

